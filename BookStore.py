import openpyxl


class Book:
    isbn = ""
    name = ""
    author = ""
    publisher = ""
    price = ""
    quantity = ""
    sold = ""
    available = ""
    info = []
    data = []

    def updateInfoData(self):
        self.info = []
        self.info.append(str(self.isbn))
        self.info.append(str(self.name))
        self.info.append(str(self.author))
        self.info.append(str(self.publisher))
        self.info.append(str(self.price))

    def updateBookData(self):
        self.data = []
        self.data.append(self.isbn)
        self.data.append(self.name)
        self.data.append(self.publisher)
        self.data.append(self.price)
        self.data.append(self.quantity)
        self.data.append(self.sold)
        self.data.append(self.available)

    def addBook(self, data):
        self.isbn = str(data[0])
        self.name = str(data[1])
        self.author = str(data[2])
        self.publisher = str(data[3])
        self.price = str(data[4])
        self.quantity = str(data[5])
        self.sold = str(data[6])
        self.available = str(data[7])
        self.updateBookData()
        self.updateInfoData()

    def addBookData(self, isbn, name, author, price, quantity, sold, available):
        self.isbn = str(isbn)
        self.name = str(name)
        self.author = str(author)
        self.price = str(price)
        self.quantity = str(quantity)
        self.sold = str(sold)
        self.available = str(available)
        self.updateBookData()
        self.updateInfoData()

    def editBook(self, isbn, name, author, price, quantity, sold, available):
        self.isbn = isbn
        self.name = name
        self.author = author
        self.price = price
        self.quantity = quantity
        self.sold = sold
        self.available = available
        self.updateInfoData()
        self.updateBookData()

    def editQuantity(self, quantity):
        self.quantity = quantity
        self.updateBookData()

    def editSold(self, sold):
        self.sold = sold
        self.updateBookData()

    def editAvailable(self, available):
        self.available = available
        self.updateBookData()


class Inventory:
    books = []
    amount_of_books = 0

    def getInventory(self):
        for row in range(1, InventorySheet.max_row+1):
            temp = []
            for col in range(1, InventorySheet.max_column+1):
                temp.append(InventorySheet.cell(row, col).value)
            self.books.append(Book())
            self.books[row-1].addBook(temp)
            self.amount_of_books += 1

    def updateInventory(self):
        for row in range(1, InventorySheet.max_row+1):
            book = self.books[row-1]
            for col in range(1, InventorySheet.max_column+1):
                for data in book.data:
                    InventorySheet.cell(row, col).value = data
        wb.save(path)

    def getPriceWithISBN(self, isbn):
        for book in books:
            if book.isbn == isbn:
                return float(book.price)
            else:
                return 0

    def bought(self, isbn):
        for book in self.books:
            if book.isbn == isbn:
                book.sold = str(float(book.sold)+1)
                book.available = str(float(book.quantity)-float(book.sold))

    def allIsbns(self):
        isbns = set()
        for book in self.books:
            if book.isbn != '':
                isbns.add(book.isbn)
        isbns = sorted(isbns)
        return list(isbns)

    def allBookNames(self):
        name = set()
        for book in self.books:
            if book.name != '':
                name.add(book.name)
        name = sorted(name)
        return list(name)

    def allAuthors(self):
        authors = set()
        for book in self.books:
            if book.author != '':
                authors.add(book.author)
        authors = sorted(authors)
        return list(authors)

    def allPublishers(self):
        publishers = set()
        for book in self.books:
            if book.publisher != '':
                publishers.add(book.publisher)
        publishers = sorted(publishers)
        return list(publishers)


class Cart(Inventory):
    cart = []
    total = 0

    def addToCart(self, isbn):
        for book in self.books:
            if book.isbn == isbn:
                self.cart.append(book)
                self.total += book.price

    def removeFromCart(self, isbn):
        for book in self.cart:
            if book.isbn == isbn:
                self.total -= book.price
                self.cart.remove(book)

    def checkOut(self):
        fw = open("bill.txt", 'w')
        for book in self.cart:
            temp = '|'.join(book.info)
            fw.write(temp + "\n")
        fw.close()
        for item in cart:
            self.bought(item[0])


path = "BookStoreData.xlsx"
wb = openpyxl.load_workbook(path)
InventorySheet = wb['Inventory']
BillsSheet = wb['Bills']
